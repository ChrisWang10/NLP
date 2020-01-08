import pickle
from openpyxl import load_workbook
import jieba
import os
import re
from tqdm import tqdm

source_path = r'C:\Users\v_wangchao3\code\MyProject\result.xlsx'
target = [r'C:\Users\v_wangchao3\code\MyProject\MyTask\text_classification\data\corpus.txt']
stop_words_path = r'C:\Users\v_wangchao3\code\MyProject\MyTask\text_classification\stopwords'
substitute = u"([^\u4e00-\u9fff\u0030-\u0039\u0041-\u005a\u0061-\u007a-\uFF00-\uFFEF-\u0030-\u0039-\uff74])"
jieba.add_word('职业发展')
jieba.add_word('恋爱关系')
jieba.add_word('心理方面')


def check_contain_chinese(check_str):
    flag = True
    for ch in check_str:
        if u'\u4e00' >= ch or ch >= u'\u9fff':
            flag = False
    return flag


def gen_stop_words_dict():
    stop_words_dict = {}
    for file in os.listdir(stop_words_path):
        with open(os.path.join(stop_words_path, file), 'r', encoding='UTF-8') as fp:
            for line in tqdm(fp.readlines()):
                stop_words_dict[line.strip()] = 1
    with open('stop_words.txt', 'wb') as fp:
        pickle.dump(stop_words_dict, fp)


def stop_words_dict():
    with open('stop_words.txt', 'rb') as fp:
        stop_words_dict = pickle.load(fp)
        return stop_words_dict


stop_words_dict = stop_words_dict()
stop_words_dict['年'] = 1
stop_words_dict['月'] = 1
stop_words_dict['日'] = 1


def remove_stop_words(source):
    ret = ''
    for words in source:
        if words in stop_words_dict:
            continue
        # if re.search('[a-z]', words) or re.search('[A-Z]', words) or re.search('[0-1]', words):
        #     continue
        if check_contain_chinese(words):
            ret += ' ' + re.sub(substitute, "", words)
    return ret


def gen_corpus():
    wb = load_workbook(source_path)
    ws1 = wb['content']

    save_file = open(target[0], 'w', encoding='utf-8')

    for j, row in tqdm(enumerate(ws1.iter_rows())):
        row_txt = ''
        for i, cell in enumerate(row):
            if type(cell.value) == str:
                row_txt += re.sub(substitute, "", cell.value)
            elif type(cell.value) == int:
                row_txt += str(cell.value)
            row_txt += ' '
        seg_list = jieba.cut(row_txt)
        result = remove_stop_words(seg_list)
        save_file.write(result + '\n')


def get_label(corpus):
    labels = open('../data/labels.txt', 'w', encoding='UTF-8')
    with open(corpus, 'r', encoding='UTF-8') as fp:
        for line in tqdm(fp.readlines()):
            label = line.strip().split(' ')[-1]
            print(label)
            labels.write(label + '\n')


if __name__ == '__main__':
    # gen_corpus()
    get_label(target[0])
